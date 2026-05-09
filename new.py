import streamlit as st
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import plotly.graph_objects as go
import plotly.express as px
import hashlib
import json
import os
import re
import smtplib
import time
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from datetime import datetime
from io import BytesIO
from pathlib import Path
import pytz
from openpyxl import load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (
    SimpleDocTemplate, 
    Table, 
    TableStyle, 
    Paragraph, 
    Spacer, 
    Image
)
import warnings

warnings.filterwarnings('ignore')

# CONFIGURATION
st.set_page_config(
    page_title="EduTrack Pro 2026",
    page_icon="🎓",
    layout="wide",
    initial_sidebar_state="expanded"
)

SMTP_CONFIG = {
    'server': 'smtp.gmail.com',
    'port': 587,
    'email': 'your_email@gmail.com',
    'password': 'your_app_password'
}


# THEME
def apply_professional_theme():
    st.markdown("""
    <style>
    .stApp > footer,
    .stApp footer,
    footer,
    div[data-testid="stApp"] footer {
        display: none !important;
        visibility: hidden !important;
        opacity: 0 !important;
        height: 0 !important;
    }
    
    #MainMenu,
    header[data-testid="stHeader"],
    div[data-testid="stToolbar"],
    div[data-testid="stDecoration"] {
        display: none !important;
    }
    
    .stApp button[title="Manage app"],
    button:has(span:contains("Manage")),
    a:has(span:contains("Manage")) {
        display: none !important;
    }
    
    div[data-testid="stBottom"],
    div[data-testid="stBottomBlock"] {
        display: none !important;
        visibility: hidden !important;
    }
    
    .main {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        font-family: 'Segoe UI', 'Inter', 'Roboto', sans-serif;
    }
    
    @keyframes gradientAnimation {
        0% { background-position: 0% 50%; }
        50% { background-position: 100% 50%; }
        100% { background-position: 0% 50%; }
    }
    
    .header {
        background: linear-gradient(-45deg, #ee7752, #e73c7e, #23a6d5, #23d5ab);
        background-size: 400% 400%;
        animation: gradientAnimation 15s ease infinite;
        color: white;
        padding: 1.8rem;
        border-radius: 0 0 25px 25px;
        margin-bottom: 2rem;
        text-align: center;
        box-shadow: 0 8px 32px rgba(31, 38, 135, 0.37);
    }
    
    .card {
        background: rgba(255, 255, 255, 0.95);
        border-radius: 20px;
        padding: 2rem;
        margin-bottom: 1.8rem;
        box-shadow: 0 8px 32px rgba(31, 38, 135, 0.15);
        transition: all 0.4s ease;
    }
    .card:hover {
        transform: translateY(-5px);
        box-shadow: 0 20px 40px rgba(31, 38, 135, 0.25);
    }
    
    .metric-card {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white;
        border-radius: 20px;
        padding: 1.5rem;
        text-align: center;
        box-shadow: 0 10px 30px rgba(102, 126, 234, 0.3);
    }
    
    .metric-value {
        font-size: 2.5rem;
        font-weight: 800;
        margin: 0.5rem 0;
    }
    
    .metric-label {
        font-size: 0.9rem;
        color: rgba(255,255,255,0.9);
        font-weight: 600;
        text-transform: uppercase;
        letter-spacing: 1px;
    }
    
    .stButton > button {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white;
        border: none;
        padding: 0.8rem 1.5rem;
        border-radius: 12px;
        font-weight: 600;
        transition: all 0.3s ease;
        width: 100%;
    }
    .stButton > button:hover {
        transform: translateY(-2px);
        box-shadow: 0 10px 25px rgba(102, 126, 234, 0.4);
    }
    </style>
    """, unsafe_allow_html=True)


# SESSION STATE INITIALIZATION
if 'logged_in' not in st.session_state: st.session_state.logged_in = False
if 'user_type' not in st.session_state: st.session_state.user_type = ""
if 'username' not in st.session_state: st.session_state.username = ""
if 'user_data' not in st.session_state: st.session_state.user_data = {}
if 'current_page' not in st.session_state: st.session_state.current_page = "login"
if 'results' not in st.session_state: st.session_state.results = {}
if 'processed' not in st.session_state: st.session_state.processed = False
if 'activity_log' not in st.session_state: st.session_state.activity_log = []
if 'teacher_uploads' not in st.session_state: st.session_state.teacher_uploads = {}
if 'show_email_modal' not in st.session_state: st.session_state.show_email_modal = None
if 'show_reset_confirm' not in st.session_state: st.session_state.show_reset_confirm = None
if 'show_change_password' not in st.session_state: st.session_state.show_change_password = False


# UTILITY FUNCTIONS
def get_current_semester():
    bd_tz = pytz.timezone('Asia/Dhaka')
    now = datetime.now(bd_tz)
    if 1 <= now.month <= 6:
        return f"Spring {now.year}"
    else:
        return f"Summer {now.year}"

def hash_password(password):
    salt = "EduTrack2026!"
    return hashlib.sha256((password + salt).encode()).hexdigest()

def verify_password(password, hashed):
    return hash_password(password) == hashed

def calculate_sgpa(total_marks):
    if total_marks >= 80: return 4.00
    elif total_marks >= 75: return 3.75
    elif total_marks >= 70: return 3.50
    elif total_marks >= 65: return 3.25
    elif total_marks >= 60: return 3.00
    elif total_marks >= 55: return 2.75
    elif total_marks >= 50: return 2.50
    elif total_marks >= 45: return 2.25
    elif total_marks >= 40: return 2.00
    else: return 0.00

def get_grade_from_marks(total_marks):
    if total_marks >= 80: return "A+"
    elif total_marks >= 75: return "A"
    elif total_marks >= 70: return "A-"
    elif total_marks >= 65: return "B+"
    elif total_marks >= 60: return "B"
    elif total_marks >= 55: return "B-"
    elif total_marks >= 50: return "C+"
    elif total_marks >= 45: return "C"
    elif total_marks >= 40: return "D"
    else: return "F"

def log_activity(username, action, details=""):
    st.session_state.activity_log.append({
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "username": username,
        "action": action,
        "details": details
    })

def save_course_data(semester, course_code, results):
    try:
        data_dir = Path("course_data")
        data_dir.mkdir(exist_ok=True)
        course_file = data_dir / f"course_{semester.replace(' ', '_')}_{course_code.replace(' ', '_')}.pkl"
        with open(course_file, 'wb') as f:
            pickle.dump(results, f)
        return True
    except Exception as e:
        st.error(f"Error saving: {e}")
        return False

def load_all_courses():
    courses = {}
    try:
        data_dir = Path("course_data")
        if data_dir.exists():
            for file in data_dir.glob("course_*.pkl"):
                try:
                    with open(file, 'rb') as f:
                        course_data = pickle.load(f)
                        key = f"{course_data.get('semester', '')} - {course_data.get('course_code', '')}"
                        courses[key] = course_data
                except:
                    continue
    except:
        pass
    return courses

def load_student_data(student_id):
    try:
        student_file = Path("course_data") / f"student_{student_id.replace(' ', '_')}.pkl"
        if student_file.exists():
            with open(student_file, 'rb') as f:
                return pickle.load(f)
        return {}
    except:
        return {}

def track_teacher_upload(username, semester, course_code, filename, student_count):
    if username not in st.session_state.teacher_uploads:
        st.session_state.teacher_uploads[username] = []
    st.session_state.teacher_uploads[username].append({
        'semester': semester,
        'course_code': course_code,
        'filename': filename,
        'student_count': student_count,
        'timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    })

def load_users():
    default_users = {
        "admins": {
            "admin": {
                "username": "admin",
                "password": hash_password("admin123"),
                "email": "admin@stamford.edu.bd",
                "full_name": "System Administrator",
                "department": "IT & Administration",
                "designation": "System Admin",
                "user_type": "admin",
                "created_at": "2024-01-01",
                "last_login": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "is_active": True,
                "admin_level": "super"
            }
        },
        "teachers": {
            "teacher": {
                "username": "teacher",
                "password": hash_password("teacher123"),
                "email": "teacher@stamford.edu.bd",
                "full_name": "Teacher",
                "department": "Electrical & Electronic Engineering",
                "designation": "lecturer",
                "user_type": "teacher",
                "created_at": "2024-01-01",
                "last_login": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "is_active": True,
                "courses_assigned": ["Power System Protection", "Circuit Theory", "Power Electronics"]
            }
        },
        "students": {},
        "parents": {}
    }
    
    try:
        if os.path.exists("users_enhanced.json"):
            with open("users_enhanced.json", 'r') as f:
                loaded_users = json.load(f)
            
            for user_type in default_users:
                if user_type not in loaded_users:
                    loaded_users[user_type] = {}
                for username, user_data in default_users[user_type].items():
                    if username not in loaded_users[user_type]:
                        loaded_users[user_type][username] = user_data
                    else:
                        if username in ['admin', 'teacher']:
                            loaded_users[user_type][username]['password'] = hash_password('admin123' if username == 'admin' else 'teacher123')
                            loaded_users[user_type][username]['is_active'] = True
            
            return loaded_users
    except Exception as e:
        st.error(f"Error loading users: {e}")
    
    return default_users

def save_users(users):
    try:
        with open("users_enhanced.json", 'w') as f:
            json.dump(users, f, indent=4, default=str)
        return True
    except:
        return False

def authenticate_user(username, password, user_type):
    users = load_users()
    user_category = user_type + "s"
    if user_category in users and username in users[user_category]:
        user_data = users[user_category][username]
        if verify_password(password, user_data["password"]):
            if not user_data.get("is_active", True):
                return False, "Account deactivated"
            return True, user_data
    return False, "Invalid credentials"

def change_password(username, user_type, old_password, new_password):
    """Change password for a user"""
    users = load_users()
    user_category = user_type + "s"
    
    if user_category in users and username in users[user_category]:
        user_data = users[user_category][username]
        
        # Verify old password
        if not verify_password(old_password, user_data["password"]):
            return False, "Current password is incorrect"
        
        # Update password
        users[user_category][username]["password"] = hash_password(new_password)
        users[user_category][username]["password_changed_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        if save_users(users):
            log_activity(username, "password_changed", f"Password changed for {username}")
            return True, "Password changed successfully!"
        
    return False, "User not found"

def send_email(to_email, subject, body, pdf_buffer=None):
    try:
        msg = MIMEMultipart()
        msg['From'] = SMTP_CONFIG['email']
        msg['To'] = to_email
        msg['Subject'] = subject
        msg.attach(MIMEText(body, 'html'))
        if pdf_buffer:
            pdf_attachment = MIMEApplication(pdf_buffer.getvalue(), _subtype='pdf')
            pdf_attachment.add_header('Content-Disposition', 'attachment', filename='Student_Report.pdf')
            msg.attach(pdf_attachment)
        with smtplib.SMTP(SMTP_CONFIG['server'], SMTP_CONFIG['port']) as server:
            server.starttls()
            server.login(SMTP_CONFIG['email'], SMTP_CONFIG['password'])
            server.send_message(msg)
        return True, "Email sent"
    except Exception as e:
        return False, str(e)


# SPIDER PLOT FUNCTION
def create_spider_plot(values, labels, title="PO Attainment Spider Plot"):
    num_vars = len(labels)
    angles = np.linspace(0, 2 * np.pi, num_vars, endpoint=False).tolist()
    values = np.concatenate((values, [values[0]]))
    angles += angles[:1]
    
    fig, ax = plt.subplots(figsize=(8, 8), subplot_kw=dict(projection='polar'))
    ax.fill(angles, values, color='#667eea', alpha=0.25)
    ax.plot(angles, values, color='#667eea', linewidth=2, marker='o', markersize=8)
    ax.set_xticks(angles[:-1])
    ax.set_xticklabels(labels, fontsize=10)
    ax.set_ylim(0, 100)
    ax.set_yticks([20, 40, 60, 80, 100])
    ax.set_yticklabels(['20%', '40%', '60%', '80%', '100%'], fontsize=8, color='gray')
    ax.grid(True, alpha=0.3)
    ax.set_title(title, size=16, fontweight='bold', pad=20)
    
    for i, (angle, value) in enumerate(zip(angles[:-1], values[:-1])):
        ax.annotate(f'{value:.1f}%', xy=(angle, value), xytext=(5, 5),
                   textcoords='offset points', fontsize=9, fontweight='bold', color='#333')
    
    plt.tight_layout()
    return fig


# ENHANCED EXCEL PARSER WITH DYNAMIC PO DETECTION
def parse_excel_eee_format(uploaded_file):
    """Enhanced parser with dynamic PO column detection - supports any number of POs"""
    try:
        wb = load_workbook(uploaded_file, data_only=True)
        sheet_names = wb.sheetnames
        
        st.info(f"📊 Found {len(sheet_names)} sheets: {', '.join(sheet_names)}")
        
        student_id_pattern = re.compile(r'EEE\s*\d{3}\s*\d{5}', re.IGNORECASE)
        co_header_pattern = re.compile(r'^CO(\d+)$', re.IGNORECASE)
        po_header_pattern = re.compile(r'^PO\(([a-zA-Z])\)$', re.IGNORECASE)  # Dynamic PO detection
        
        parsed_data = {
            'course_info': {
                'university': 'Stamford University Bangladesh',
                'trimester': '',
                'section': '',
                'course_code': '',
                'course_title': '',
                'teacher': ''
            },
            'students': {},
            'co_columns': [],
            'po_columns': [],
            'co_po_mapping': {},
            'max_co_marks': {},
            'max_po_marks': {},
        }
        
        # Parse Course Info from Midterm Exam sheet
        if 'Midterm Exam' in sheet_names:
            mid_sheet = wb['Midterm Exam']
            for row in mid_sheet.iter_rows(min_row=1, max_row=8, max_col=4, values_only=True):
                for cell_val in row:
                    if cell_val and isinstance(cell_val, str):
                        cell_str = str(cell_val).strip()
                        if 'Trimester' in cell_str:
                            parsed_data['course_info']['trimester'] = cell_str.split(':')[-1].strip()
                        elif 'Section' in cell_str and 'Course' not in cell_str:
                            parsed_data['course_info']['section'] = cell_str.split(':')[-1].strip()
                        elif 'Course Code' in cell_str:
                            parsed_data['course_info']['course_code'] = cell_str.split(':')[-1].strip()
                        elif 'Course Title' in cell_str:
                            parsed_data['course_info']['course_title'] = cell_str.split(':')[-1].strip()
                        elif 'Course Teacher' in cell_str or 'Teacher' in cell_str:
                            parsed_data['course_info']['teacher'] = cell_str.split(':')[-1].strip().replace(':', '').strip()
        
        # Parse Analysis of CO sheet for student data
        if 'Analysis of CO' in sheet_names:
            co_sheet = wb['Analysis of CO']
            
            # Find the header row (row with "SL" in column A)
            sl_header_row = None
            for row_idx in range(1, min(30, co_sheet.max_row + 1)):
                cell_val = str(co_sheet.cell(row=row_idx, column=1).value or '').strip()
                if cell_val.upper() == 'SL':
                    sl_header_row = row_idx
                    break
            
            if sl_header_row:
                # Find CO columns dynamically
                co_columns_found = []
                for col_idx in range(5, 20):  # Scan columns E to T
                    cell_val = str(co_sheet.cell(row=sl_header_row, column=col_idx).value or '').strip()
                    co_match = co_header_pattern.match(cell_val)
                    if co_match:
                        co_name = f"CO{co_match.group(1)}"
                        co_columns_found.append((col_idx, co_name))
                
                parsed_data['co_columns'] = [co_name for _, co_name in co_columns_found]
                
                # Max marks row is one row below header
                max_row = sl_header_row + 1
                
                # Parse max CO marks
                for col_idx, co_name in co_columns_found:
                    val = co_sheet.cell(row=max_row, column=col_idx).value
                    parsed_data['max_co_marks'][co_name] = float(val) if val else 0
                
                # Parse individual student data (rows after max_row)
                for row_idx in range(max_row + 1, co_sheet.max_row + 1):
                    sl_val = co_sheet.cell(row=row_idx, column=1).value
                    if not sl_val:
                        continue
                    
                    # Student ID is in column B (index 2)
                    student_id_cell = co_sheet.cell(row=row_idx, column=2).value
                    if not student_id_cell:
                        continue
                    
                    student_id = str(student_id_cell).strip()
                    
                    # Validate student ID format
                    if not student_id_pattern.search(student_id):
                        if not re.match(r'^\d{3}\s*\d{5}$', student_id) and not re.match(r'^EEE', student_id, re.IGNORECASE):
                            continue
                    
                    student_name = str(co_sheet.cell(row=row_idx, column=3).value or '').strip()
                    student_status = str(co_sheet.cell(row=row_idx, column=4).value or '').strip()
                    
                    # Read CO marks from dynamically found columns
                    co_marks = {}
                    co_pct = {}
                    
                    for col_idx, co_name in co_columns_found:
                        val = co_sheet.cell(row=row_idx, column=col_idx).value
                        co_marks[co_name] = float(val) if val is not None else 0.0
                        
                        # CO attainment percentages are usually 2 columns after marks
                        pct_col = col_idx + 2
                        pct_val = co_sheet.cell(row=row_idx, column=pct_col).value
                        co_pct[co_name] = float(pct_val) if pct_val is not None else 0.0
                    
                    total_marks = sum(co_marks.values())
                    
                    parsed_data['students'][student_id] = {
                        'id': student_id,
                        'name': student_name,
                        'status': student_status,
                        'co_marks': co_marks,
                        'total_marks': total_marks,
                        'co_attainment_pct': co_pct,
                        'po_marks': {},
                        'po_attainment_pct': {},
                        'sgpa': calculate_sgpa(total_marks),
                        'grade': get_grade_from_marks(total_marks),
                        'status_final': 'Pass' if total_marks >= 40 else 'Fail',
                    }
        
        # Parse Analysis of PO sheet with DYNAMIC PO detection
        if 'Analysis of PO' in sheet_names:
            po_sheet = wb['Analysis of PO']
            
            # Find CO-PO matrix header
            co_po_start = None
            for row_idx in range(1, min(25, po_sheet.max_row + 1)):
                cell_val = str(po_sheet.cell(row=row_idx, column=3).value or '').strip()
                if 'CO-PO matrix' in cell_val or 'CO-PO' in cell_val:
                    co_po_start = row_idx
                    break
            
            if not co_po_start:
                co_po_start = 10
            
            # DYNAMICALLY detect PO columns from the CO-PO matrix row
            po_columns = {}
            for col_idx in range(4, 30):  # Scan columns D to AD (wide range)
                cell_val = str(po_sheet.cell(row=co_po_start, column=col_idx).value or '').strip()
                po_match = po_header_pattern.match(cell_val)
                if po_match:
                    po_name = f"PO({po_match.group(1).lower()})"
                    po_columns[col_idx] = po_name
            
            parsed_data['po_columns'] = sorted(po_columns.values())
            st.info(f"📊 Dynamically detected {len(parsed_data['po_columns'])} POs: {', '.join(parsed_data['po_columns'])}")
            
            # Parse CO-PO mapping weights
            co_po_mapping = {}
            for row_offset in range(1, len(parsed_data['co_columns']) + 1):
                row_idx = co_po_start + row_offset
                co_cell = str(po_sheet.cell(row=row_idx, column=3).value or '').strip()
                co_match = co_header_pattern.match(co_cell)
                if co_match:
                    co_name = f"CO{co_match.group(1)}"
                    mapping = {}
                    for col_idx, po_name in po_columns.items():
                        val = po_sheet.cell(row=row_idx, column=col_idx).value
                        if val and float(val) > 0:
                            mapping[po_name] = float(val)
                    if mapping:
                        co_po_mapping[co_name] = mapping
            
            parsed_data['co_po_mapping'] = co_po_mapping
            
            # Find student data section in PO sheet
            student_header_row = None
            for row_idx in range(25, min(50, po_sheet.max_row + 1)):
                cell_val = str(po_sheet.cell(row=row_idx, column=1).value or '').strip()
                if cell_val.upper() == 'SL':
                    student_header_row = row_idx
                    break
            
            if student_header_row:
                max_po_row = student_header_row + 1
                
                # Parse PO max marks from the row after header
                for col_idx, po_name in po_columns.items():
                    val = po_sheet.cell(row=max_po_row, column=col_idx).value
                    parsed_data['max_po_marks'][po_name] = float(val) if val else 0
                
                # Parse individual student PO data
                for row_idx in range(max_po_row + 1, po_sheet.max_row + 1):
                    sl_val = po_sheet.cell(row=row_idx, column=1).value
                    if not sl_val:
                        continue
                    
                    student_id_cell = po_sheet.cell(row=row_idx, column=2).value
                    if not student_id_cell:
                        continue
                    
                    student_id = str(student_id_cell).strip()
                    if not student_id_pattern.search(student_id):
                        if not re.match(r'^\d{3}\s*\d{5}$', student_id) and not re.match(r'^EEE', student_id, re.IGNORECASE):
                            continue
                    
                    if student_id not in parsed_data['students']:
                        continue
                    
                    # Read PO marks from dynamically detected columns
                    po_marks = {}
                    po_pct = {}
                    
                    for col_idx, po_name in po_columns.items():
                        # PO marks
                        mark_val = po_sheet.cell(row=row_idx, column=col_idx).value
                        po_marks[po_name] = float(mark_val) if mark_val is not None else 0.0
                        
                        # PO percentages (usually 2 columns after marks, but find dynamically)
                        # Look for percentage in nearby columns
                        for pct_offset in [2, 3, 4]:
                            pct_col = col_idx + pct_offset
                            if pct_col <= po_sheet.max_column:
                                pct_val = po_sheet.cell(row=row_idx, column=pct_col).value
                                if pct_val and isinstance(pct_val, (int, float)) and 0 <= pct_val <= 100:
                                    po_pct[po_name] = float(pct_val)
                                    break
                    
                    parsed_data['students'][student_id]['po_marks'] = po_marks
                    parsed_data['students'][student_id]['po_attainment_pct'] = po_pct
        
        wb.close()
        
        student_count = len(parsed_data['students'])
        if student_count > 0:
            st.success(f"✅ Successfully parsed {student_count} students with dynamic PO detection")
            with st.expander("📋 Parsing Details"):
                st.write(f"**CO columns:** {parsed_data['co_columns']}")
                st.write(f"**PO columns (auto-detected):** {parsed_data['po_columns']}")
                st.write(f"**CO-PO mappings:** {len(parsed_data['co_po_mapping'])}")
                st.write(f"**Course Info:** {parsed_data['course_info']}")
                
                # Show sample data
                sample_students = list(parsed_data['students'].items())[:3]
                if sample_students:
                    sample_data = []
                    for sid, s in sample_students:
                        sample_data.append({
                            'Student': s['name'][:20],
                            'Total Marks': f"{s['total_marks']:.1f}",
                            'COs': len(s.get('co_marks', {})),
                            'POs': len(s.get('po_marks', {}))
                        })
                    st.dataframe(pd.DataFrame(sample_data), use_container_width=True)
        else:
            st.warning("⚠️ No students found. Please check the Excel file format.")
        
        return parsed_data
        
    except Exception as e:
        st.error(f"❌ Error parsing Excel: {str(e)}")
        import traceback
        st.code(traceback.format_exc())
        return None


# CONVERT PARSED DATA
def convert_parsed_to_results(parsed_data, semester, course_code, course_name, username):
    """Convert parsed data to standard results format"""
    results = {
        'students': {},
        'course_stats': {},
        'semester': semester,
        'course_code': course_code,
        'course_name': course_name or parsed_data.get('course_info', {}).get('course_title', ''),
        'teacher': username,
        'co_columns': parsed_data.get('co_columns', []),
        'po_columns': parsed_data.get('po_columns', []),
        'co_po_mapping': parsed_data.get('co_po_mapping', {}),
        'timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    }
    
    co_po_mapping = results['co_po_mapping']
    
    # Calculate course-level CO attainment (average of all students)
    co_attainment = {}
    for co in results['co_columns']:
        scores = [s['co_attainment_pct'].get(co, 0) for s in parsed_data['students'].values()]
        if scores:
            co_attainment[co] = round(np.mean(scores), 2)
    results['co_attainment'] = co_attainment
    
    # Calculate course-level PO attainment using CO-PO mapping
    po_attainment = {}
    if co_po_mapping and co_attainment:
        for po in results['po_columns']:
            po_value = 0
            for co, weights in co_po_mapping.items():
                if po in weights:
                    po_value += co_attainment.get(co, 0) * weights[po]
            po_attainment[po] = round(min(100, po_value), 2)
    results['po_attainment'] = po_attainment
    
    # Individual student records
    for student_id, student in parsed_data['students'].items():
        # Use individual CO attainment percentages as co_scores
        co_scores = {}
        for co, pct in student.get('co_attainment_pct', {}).items():
            co_scores[co] = round(pct * 20 / 100, 2)  # Convert % to score out of 20
        
        # Calculate individual PO attainment using individual CO percentages
        individual_po = {}
        if co_po_mapping and student.get('co_attainment_pct'):
            for po in results['po_columns']:
                po_value = 0
                for co, weights in co_po_mapping.items():
                    if po in weights:
                        student_co_pct = student['co_attainment_pct'].get(co, 0)
                        po_value += student_co_pct * weights[po]
                individual_po[po] = round(min(100, po_value), 2)
        
        # Use parsed PO attainment if available
        if student.get('po_attainment_pct'):
            individual_po.update(student['po_attainment_pct'])
        
        results['students'][student_id] = {
            'id': student_id,
            'name': student.get('name', ''),
            'mid': 0,
            'final': 0,
            'ct': 0,
            'assignment': 0,
            'attendance': 0,
            'academic_total': student.get('total_marks', 0),
            'total_marks': student.get('total_marks', 0),
            'sgpa': student.get('sgpa', 0),
            'grade': student.get('grade', 'F'),
            'co_scores': co_scores,
            'co_attainment_pct': student.get('co_attainment_pct', {}),
            'po_scores': individual_po,
            'student_email': '',
            'parent_email': '',
            'course_code': course_code,
            'semester': semester,
            'status': student.get('status_final', 'Fail'),
            'timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }
    
    # Calculate course statistics
    if results['students']:
        marks = [s['total_marks'] for s in results['students'].values()]
        passing = len([m for m in marks if m >= 40])
        total = len(marks)
        results['course_stats'] = {
            'average_marks': round(np.mean(marks), 2),
            'academic_average': round(np.mean(marks), 2),
            'highest_marks': round(max(marks), 2),
            'lowest_marks': round(min(marks), 2),
            'average_sgpa': round(np.mean([s['sgpa'] for s in results['students'].values()]), 2),
            'total_students': total,
            'passing_students': passing,
            'pass_percentage': round((passing / total * 100) if total > 0 else 0, 1),
            'fail_percentage': round(((total - passing) / total * 100) if total > 0 else 0, 1),
            'std_deviation': round(np.std(marks), 2) if marks else 0
        }
    
    return results


# COMPREHENSIVE PDF REPORT
def generate_comprehensive_pdf_report(course_data, course_name):
    """Generate comprehensive PDF report in 2 pages"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=30, leftMargin=30, topMargin=25, bottomMargin=20)
    
    styles = getSampleStyleSheet()
    story = []
    
    # Custom styles
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=14, alignment=1, spaceAfter=8, textColor=colors.HexColor('#1a237e'))
    heading_style = ParagraphStyle('Heading', parent=styles['Heading2'], fontSize=11, spaceAfter=5, spaceBefore=5, textColor=colors.HexColor('#283593'))
    normal_style = ParagraphStyle('Normal', parent=styles['Normal'], fontSize=8, spaceAfter=3)
    
    # PAGE 1 HEADER
    semester = course_data.get('semester', 'N/A')
    course_code = course_data.get('course_code', 'N/A')
    course_name_display = course_data.get('course_name', course_name)
    teacher = course_data.get('teacher', 'N/A')
    
    story.append(Paragraph("EduTrack Pro 2026 - Comprehensive Course Report", title_style))
    story.append(Spacer(1, 5))
    
    # Info header
    header_data = [
        [Paragraph(f"<b>Course:</b> {course_code} - {course_name_display}", normal_style),
         Paragraph(f"<b>Semester:</b> {semester}", normal_style)],
        [Paragraph(f"<b>Teacher:</b> {teacher}", normal_style),
         Paragraph(f"<b>Generated:</b> {datetime.now().strftime('%Y-%m-%d %H:%M')}", normal_style)]
    ]
    header_table = Table(header_data, colWidths=[280, 250])
    header_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#e8eaf6')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#c5cae9')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    story.append(header_table)
    story.append(Spacer(1, 8))
    
    # COURSE STATISTICS
    story.append(Paragraph("Course Statistics", heading_style))
    
    stats = course_data.get('course_stats', {})
    stats_data = [
        ['Total Students', str(stats.get('total_students', 0)),
         'Average Marks', f"{stats.get('average_marks', 0):.1f}"],
        ['Highest Marks', f"{stats.get('highest_marks', 0):.1f}",
         'Lowest Marks', f"{stats.get('lowest_marks', 0):.1f}"],
        ['Pass Rate', f"{stats.get('pass_percentage', 0):.1f}%",
         'Avg SGPA', f"{stats.get('average_sgpa', 0):.2f}"],
        ['Std Deviation', f"{stats.get('std_deviation', 0):.2f}",
         'Grade Range', f"A+ to F"],
    ]
    
    stats_table = Table(stats_data, colWidths=[100, 130, 100, 130])
    stats_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#1a237e')),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.whitesmoke),
        ('BACKGROUND', (2, 0), (2, -1), colors.HexColor('#1a237e')),
        ('TEXTCOLOR', (2, 0), (2, -1), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#bdbdbd')),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]))
    story.append(stats_table)
    story.append(Spacer(1, 8))
    
    # MARKS & GRADE DISTRIBUTION
    students = course_data.get('students', {})
    if students:
        story.append(Paragraph("Performance Analytics", heading_style))
        
        marks = [s.get('total_marks', 0) for s in students.values()]
        grades = [s.get('grade', 'F') for s in students.values()]
        
        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(8.5, 3.2))
        
        # Marks Distribution Histogram
        ax1.hist(marks, bins=8, color='#667eea', edgecolor='white', alpha=0.85, linewidth=0.5)
        ax1.set_title('Marks Distribution', fontsize=10, fontweight='bold', color='#1a237e')
        ax1.set_xlabel('Total Marks', fontsize=8)
        ax1.set_ylabel('Students', fontsize=8)
        ax1.grid(True, alpha=0.2, linestyle='--')
        ax1.tick_params(labelsize=7)
        ax1.spines['top'].set_visible(False)
        ax1.spines['right'].set_visible(False)
        
        # Grade Distribution Pie
        grade_counts = pd.Series(grades).value_counts()
        grade_colors = {
            'A+': '#1B5E20', 'A': '#4CAF50', 'A-': '#8BC34A', 
            'B+': '#CDDC39', 'B': '#FFC107', 'B-': '#FF9800', 
            'C+': '#FF5722', 'C': '#9C27B0', 'D': '#673AB7', 'F': '#F44336'
        }
        colors_list = [grade_colors.get(g, '#999999') for g in grade_counts.index]
        wedges, texts, autotexts = ax2.pie(
            grade_counts.values, labels=grade_counts.index, 
            colors=colors_list, autopct='%1.1f%%', 
            startangle=90, pctdistance=0.75
        )
        for at in autotexts: 
            at.set_color('white')
            at.set_fontweight('bold')
            at.set_fontsize(6)
        for t in texts: 
            t.set_fontsize(7)
        ax2.set_title('Grade Distribution', fontsize=10, fontweight='bold', color='#1a237e')
        
        plt.tight_layout(pad=1.5)
        marks_buf = BytesIO()
        plt.savefig(marks_buf, format='png', dpi=130, bbox_inches='tight', facecolor='white')
        plt.close()
        marks_buf.seek(0)
        story.append(Image(marks_buf, width=7.5*inch, height=2.8*inch))
        
        story.append(PageBreak())
        
        # PAGE 2 - CO-PO ATTAINMENT
        story.append(Paragraph("CO-PO Attainment Analysis", heading_style))
        
        co_attainment = course_data.get('co_attainment', {})
        po_attainment = course_data.get('po_attainment', {})
        
        if co_attainment or po_attainment:
            has_co = bool(co_attainment)
            has_po = bool(po_attainment)
            
            if has_co and has_po:
                fig2, (ax_co, ax_po) = plt.subplots(1, 2, figsize=(8.5, 3.2))
                
                # CO Attainment Bar Chart
                cos = list(co_attainment.keys())
                co_vals = list(co_attainment.values())
                co_colors = ['#4CAF50' if v >= 80 else '#FFC107' if v >= 60 else '#F44336' for v in co_vals]
                bars = ax_co.bar(cos, co_vals, color=co_colors, edgecolor='white', linewidth=0.5, alpha=0.85)
                for bar, val in zip(bars, co_vals):
                    ax_co.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 1.5, 
                              f'{val:.1f}%', ha='center', fontsize=7, fontweight='bold', color='#333')
                ax_co.set_title('CO Attainment (Course Avg)', fontsize=10, fontweight='bold', color='#1a237e')
                ax_co.set_ylabel('Percentage (%)', fontsize=8)
                ax_co.set_ylim(0, max(co_vals) * 1.2 + 5 if co_vals else 105)
                ax_co.grid(True, alpha=0.2, linestyle='--', axis='y')
                ax_co.tick_params(labelsize=7)
                ax_co.spines['top'].set_visible(False)
                ax_co.spines['right'].set_visible(False)
                
                # PO Spider Plot
                pos = list(po_attainment.keys())
                po_vals = list(po_attainment.values())
                num_vars = len(pos)
                angles = np.linspace(0, 2 * np.pi, num_vars, endpoint=False).tolist()
                po_vals_plot = po_vals + [po_vals[0]]
                angles += angles[:1]
                
                ax_po.remove()
                ax_po = fig2.add_subplot(1, 2, 2, projection='polar')
                ax_po.fill(angles, po_vals_plot, color='#667eea', alpha=0.2)
                ax_po.plot(angles, po_vals_plot, color='#667eea', linewidth=1.5, marker='o', markersize=5, markerfacecolor='#764ba2')
                ax_po.set_xticks(angles[:-1])
                ax_po.set_xticklabels(pos, fontsize=7)
                ax_po.set_ylim(0, 100)
                ax_po.set_yticks([20, 40, 60, 80, 100])
                ax_po.set_yticklabels(['20%', '40%', '60%', '80%', '100%'], fontsize=6, color='gray')
                ax_po.grid(True, alpha=0.3, linestyle='--')
                ax_po.set_title('PO Attainment (Course Avg)', fontsize=10, fontweight='bold', color='#1a237e', pad=15)
                
                for i, (angle, value) in enumerate(zip(angles[:-1], po_vals_plot[:-1])):
                    ax_po.annotate(f'{value:.1f}%', xy=(angle, value), xytext=(4, 4),
                                  textcoords='offset points', fontsize=6, fontweight='bold', color='#333')
                
                plt.tight_layout(pad=1.5)
                copo_buf = BytesIO()
                plt.savefig(copo_buf, format='png', dpi=130, bbox_inches='tight', facecolor='white')
                plt.close()
                copo_buf.seek(0)
                story.append(Image(copo_buf, width=7.5*inch, height=2.8*inch))
        
        # STUDENT LIST
        story.append(Spacer(1, 10))
        story.append(Paragraph("Student Performance Summary", heading_style))
        
        sorted_students = sorted(students.items(), key=lambda x: x[1].get('total_marks', 0), reverse=True)
        
        student_table_data = [['#', 'Student ID', 'Name', 'Marks', 'Grade', 'SGPA', 'Status']]
        
        for rank, (sid, s) in enumerate(sorted_students[:30], 1):  # Limit to 30 students for PDF
            name = s.get('name', 'N/A')[:22]
            student_table_data.append([
                str(rank), sid, name,
                f"{s.get('total_marks', 0):.1f}",
                s.get('grade', 'N/A'),
                f"{s.get('sgpa', 0):.2f}",
                s.get('status', 'N/A')
            ])
        
        if len(sorted_students) > 30:
            student_table_data.append(['...', '...', f'... and {len(sorted_students) - 30} more', '...', '...', '...', '...'])
        
        student_table = Table(student_table_data, colWidths=[20, 90, 160, 45, 38, 38, 50])
        student_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a237e')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('ALIGN', (3, 1), (5, -1), 'CENTER'),
            ('ALIGN', (6, 1), (6, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 7),
            ('FONTSIZE', (0, 1), (-1, -1), 6.5),
            ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#bdbdbd')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#f5f5f5'), colors.white]),
            ('TOPPADDING', (0, 0), (-1, -1), 2),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
            ('LEFTPADDING', (0, 0), (-1, -1), 3),
            ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ]))
        story.append(student_table)
    
    # FOOTER
    story.append(Spacer(1, 15))
    story.append(Paragraph(
        f"Generated by EduTrack Pro 2026 | Stamford University Bangladesh | {datetime.now().strftime('%Y-%m-%d %H:%M')} | Page 2 of 2",
        ParagraphStyle('Footer', parent=styles['Normal'], fontSize=6, alignment=1, textColor=colors.gray)
    ))
    
    doc.build(story)
    buffer.seek(0)
    return buffer


# CHANGE PASSWORD PAGE
def show_change_password_page():
    """Allow logged-in users to change their password"""
    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown("## 🔐 Change Password")
    
    with st.form("change_password_form"):
        current_password = st.text_input("Current Password", type="password", placeholder="Enter your current password")
        new_password = st.text_input("New Password", type="password", placeholder="Enter new password")
        confirm_password = st.text_input("Confirm New Password", type="password", placeholder="Confirm new password")
        
        col1, col2 = st.columns(2)
        with col1:
            if st.form_submit_button("✅ Update Password", use_container_width=True):
                if not current_password or not new_password:
                    st.error("Please fill all fields")
                elif new_password != confirm_password:
                    st.error("New password and confirmation do not match")
                elif len(new_password) < 4:
                    st.error("Password must be at least 4 characters")
                else:
                    success, message = change_password(
                        st.session_state.username,
                        st.session_state.user_type,
                        current_password,
                        new_password
                    )
                    if success:
                        st.success(message)
                        st.balloons()
                        st.session_state.show_change_password = False
                        time.sleep(1)
                        st.rerun()
                    else:
                        st.error(message)
        
        with col2:
            if st.form_submit_button("❌ Cancel", use_container_width=True):
                st.session_state.show_change_password = False
                st.rerun()
    
    st.markdown('</div>', unsafe_allow_html=True)


# PAGES
def show_login_page():
    apply_professional_theme()
    
    st.markdown("""
    <div style="text-align: center; margin: 3rem 0;">
        <h1 style="font-size: 4rem;">🎓</h1>
        <h2>EduTrack Pro 2026</h2>
        <p style="color: #666;">Academic Analytics & Management System</p>
        <p style="color: #888; font-size: 0.9rem;">Department of EEE, Stamford University Bangladesh</p>
    </div>
    """, unsafe_allow_html=True)
    
    col1, col2, col3 = st.columns([1, 1, 1])
    
    with col2:
        with st.form("login_form"):
            st.markdown("### 🔐 Login")
            user_type = st.selectbox("Account Type", ["admin", "teacher", "student", "parent"])
            username = st.text_input("Username")
            password = st.text_input("Password", type="password")
            
            if st.form_submit_button("Login", use_container_width=True):
                if username and password:
                    success, user_data = authenticate_user(username, password, user_type)
                    if success:
                        st.session_state.logged_in = True
                        st.session_state.user_type = user_type
                        st.session_state.username = username
                        st.session_state.user_data = user_data
                        log_activity(username, "login")
                        st.session_state.current_page = "upload" if user_type in ["teacher", "admin"] else "student_analytics"
                        st.rerun()
                    else:
                        st.error("Invalid credentials")
                else:
                    st.warning("Please enter credentials")


def show_upload_page():
    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown("## 📤 Upload Student Data")
    st.markdown("Upload your Excel file with Midterm, Final, Assignment, CO, and PO analysis sheets")
    
    col1, col2, col3 = st.columns(3)
    with col1:
        semester = st.text_input("Semester*", value=get_current_semester())
    with col2:
        course_code = st.text_input("Course Code*", value="EEE 321")
    with col3:
        course_name = st.text_input("Course Name*", value="Power System I")
    
    uploaded_file = st.file_uploader(
        "Choose Excel file (.xlsx or .xls)",
        type=['xlsx', 'xls'],
        help="Upload your Excel file with sheets: Midterm Exam, Final Exam, Assignment, Analysis of CO, Analysis of PO"
    )
    
    if uploaded_file:
        st.success(f"📁 File uploaded: {uploaded_file.name}")
        
        with st.expander("📄 File Preview", expanded=False):
            try:
                df_preview = pd.read_excel(uploaded_file, sheet_name=0, header=None, nrows=25)
                st.dataframe(df_preview, use_container_width=True)
            except:
                st.warning("Could not preview file")
        
        if st.button("🚀 Process & Analyze Data", type="primary", use_container_width=True):
            if not semester or not course_code:
                st.error("Please fill semester and course code")
            else:
                with st.spinner("🔍 Analyzing file structure..."):
                    parsed = parse_excel_eee_format(uploaded_file)
                    
                    if parsed and parsed.get('students'):
                        st.success(f"✅ Successfully parsed! Found {len(parsed.get('students', {}))} students")
                        
                        with st.expander("📊 Parsing Results", expanded=True):
                            col_a, col_b, col_c = st.columns(3)
                            with col_a:
                                st.metric("Students", len(parsed.get('students', {})))
                            with col_b:
                                cos = parsed.get('co_columns', [])
                                st.metric("COs", len(cos))
                            with col_c:
                                pos = parsed.get('po_columns', [])
                                st.metric("POs (Auto-detected)", len(pos))
                        
                        results = convert_parsed_to_results(
                            parsed, semester, course_code, course_name, st.session_state.username
                        )
                        
                        st.session_state.results = results
                        st.session_state.processed = True
                        
                        # Save course data
                        import pickle
                        save_course_data(semester, course_code, results)
                        track_teacher_upload(
                            st.session_state.username, semester, course_code,
                            uploaded_file.name, len(results['students'])
                        )
                        
                        # Save individual student data
                        for student_id, student_data in results['students'].items():
                            student_file = Path("course_data") / f"student_{student_id.replace(' ', '_')}.pkl"
                            existing = {}
                            if student_file.exists():
                                with open(student_file, 'rb') as f:
                                    existing = pickle.load(f)
                            
                            existing[f"{semester}_{course_code}"] = {
                                'course_code': course_code,
                                'semester': semester,
                                'student_data': student_data,
                                'co_attainment': results.get('co_attainment', {}),
                                'po_attainment': student_data.get('po_scores', results.get('po_attainment', {})),
                                'co_po_mapping': parsed.get('co_po_mapping', {}),
                            }
                            
                            with open(student_file, 'wb') as f:
                                pickle.dump(existing, f)
                        
                        stats = results.get('course_stats', {})
                        col_m1, col_m2, col_m3, col_m4 = st.columns(4)
                        with col_m1:
                            st.metric("Students", stats.get('total_students', 0))
                        with col_m2:
                            st.metric("Avg Marks", f"{stats.get('average_marks', 0):.1f}")
                        with col_m3:
                            st.metric("Pass Rate", f"{stats.get('pass_percentage', 0):.1f}%")
                        with col_m4:
                            st.metric("Avg SGPA", f"{stats.get('average_sgpa', 0):.2f}")
                        
                        st.success("✅ Data processed successfully!")
                    else:
                        st.error("❌ Could not parse file. Please check the format matches EEE department template")
    
    st.markdown('</div>', unsafe_allow_html=True)


def show_course_reports():
    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown("## 📊 Course Reports & Analytics")
    
    all_courses = load_all_courses()
    
    if not all_courses:
        st.info("ℹ️ No course data available. Please upload data first.")
        if st.button("Go to Upload Page", use_container_width=True):
            st.session_state.current_page = "upload"
            st.rerun()
        st.markdown("</div>", unsafe_allow_html=True)
        return
    
    # Filter courses for teacher (only show their own uploads)
    if st.session_state.user_type == "teacher":
        teacher_courses = {
            k: v for k, v in all_courses.items() 
            if v.get('teacher') == st.session_state.username
        }
        all_courses = teacher_courses
        
        if not teacher_courses:
            st.info("ℹ️ You haven't uploaded any courses yet.")
            if st.button("Upload Course Data", use_container_width=True):
                st.session_state.current_page = "upload"
                st.rerun()
            st.markdown("</div>", unsafe_allow_html=True)
            return
    
    selected_course = st.selectbox("Select Course", list(all_courses.keys()))
    
    if selected_course:
        course_data = all_courses[selected_course]
        stats = course_data.get('course_stats', {})
        students = course_data.get('students', {})
        
        # ACTION BUTTONS
        st.markdown("#### 📋 Actions")
        col_action1, col_action2 = st.columns(2)
        
        with col_action1:
            if st.button("📄 Download Comprehensive PDF Report", use_container_width=True, type="primary"):
                with st.spinner("Generating comprehensive PDF report..."):
                    pdf_buffer = generate_comprehensive_pdf_report(course_data, selected_course)
                    
                    st.download_button(
                        label="⬇️ Click to Download PDF Report",
                        data=pdf_buffer,
                        file_name=f"Course_Report_{course_data.get('course_code', 'COURSE')}_{course_data.get('semester', 'SEM').replace(' ', '_')}.pdf",
                        mime="application/pdf",
                        use_container_width=True,
                        key="download_pdf_report"
                    )
                    st.success("✅ PDF report generated successfully!")
        
        with col_action2:
            users = load_users()
            valid_emails = 0
            for student_id in students.keys():
                for s_info in users.get('students', {}).values():
                    if s_info.get('student_id') == student_id:
                        if s_info.get('parent_email'):
                            valid_emails += 1
                        break
            
            if st.button(f"📧 Send Bulk Emails to Parents ({valid_emails} recipients)", use_container_width=True):
                st.session_state.show_email_modal = selected_course
                st.rerun()
        
        # Bulk email modal
        if st.session_state.get('show_email_modal') == selected_course:
            st.markdown("---")
            st.markdown("### 📧 Send Results to Parents")
            
            users = load_users()
            email_recipients = []
            
            for student_id, student_data in students.items():
                parent_email = None
                parent_name = "Parent"
                
                for s_username, s_info in users.get('students', {}).items():
                    if s_info.get('student_id') == student_id:
                        parent_email = s_info.get('parent_email', '')
                        parent_name = s_info.get('parent_name', 'Parent')
                        break
                
                if not parent_email:
                    for p_username, p_info in users.get('parents', {}).items():
                        if p_info.get('student_linked') == student_id:
                            parent_email = p_info.get('email', '')
                            parent_name = p_info.get('full_name', 'Parent')
                            break
                
                if parent_email and parent_email not in ['', 'nan', 'NaN']:
                    email_recipients.append({
                        'student_id': student_id,
                        'student_name': student_data.get('name', 'N/A'),
                        'parent_name': parent_name,
                        'parent_email': parent_email,
                        'grade': student_data.get('grade', 'N/A'),
                        'marks': student_data.get('total_marks', 0)
                    })
            
            if email_recipients:
                st.info(f"📧 Found **{len(email_recipients)}** parents with valid emails")
                
                preview_data = []
                for r in email_recipients[:10]:
                    preview_data.append({
                        'Student': r['student_name'],
                        'Parent': r['parent_name'],
                        'Email': r['parent_email'],
                        'Grade': r['grade']
                    })
                st.dataframe(pd.DataFrame(preview_data), use_container_width=True)
                
                if len(email_recipients) > 10:
                    st.caption(f"...and {len(email_recipients) - 10} more")
                
                col_confirm1, col_confirm2 = st.columns(2)
                with col_confirm1:
                    if st.button("📤 Send Bulk Emails Now", type="primary", use_container_width=True):
                        progress_bar = st.progress(0)
                        status_text = st.empty()
                        success_count = 0
                        fail_count = 0
                        
                        for idx, recipient in enumerate(email_recipients):
                            status_text.text(f"Sending: {recipient['student_name']} ({idx+1}/{len(email_recipients)})")
                            
                            body = f"""
                            <html>
                            <body style="font-family: Arial, sans-serif;">
                                <div style="max-width: 600px; margin: 0 auto; border: 1px solid #ddd; border-radius: 10px;">
                                    <div style="background: linear-gradient(135deg, #667eea, #764ba2); padding: 20px; text-align: center;">
                                        <h2 style="color: white;">🎓 EduTrack Pro 2026</h2>
                                        <p style="color: white;">Academic Result</p>
                                    </div>
                                    <div style="padding: 20px;">
                                        <p>Dear <strong>{recipient['parent_name']}</strong>,</p>
                                        <p>Result for <strong>{recipient['student_name']}</strong>:</p>
                                        <p>📊 Marks: <strong>{recipient['marks']:.1f}/100</strong></p>
                                        <p>🏆 Grade: <strong style="font-size: 20px;">{recipient['grade']}</strong></p>
                                        <p style="color: #666;">📚 Course: {selected_course}</p>
                                        <hr>
                                        <p style="font-size: 12px; color: #999;">EduTrack Pro 2026 - Stamford University Bangladesh</p>
                                    </div>
                                </div>
                            </body>
                            </html>
                            """
                            
                            success, _ = send_email(recipient['parent_email'], 
                                                   f"Result - {recipient['student_name']}", body)
                            if success:
                                success_count += 1
                            else:
                                fail_count += 1
                            
                            progress_bar.progress((idx + 1) / len(email_recipients))
                        
                        status_text.empty()
                        if success_count > 0:
                            st.success(f"✅ Sent {success_count} emails!")
                        if fail_count > 0:
                            st.error(f"❌ Failed: {fail_count}")
                        
                        st.session_state.show_email_modal = None
                        log_activity(st.session_state.username, "bulk_email", f"Sent {success_count} emails for {selected_course}")
                
                with col_confirm2:
                    if st.button("❌ Cancel", use_container_width=True):
                        st.session_state.show_email_modal = None
                        st.rerun()
            else:
                st.warning("⚠️ No parent emails found. Please register parents first.")
                if st.button("Close", use_container_width=True):
                    st.session_state.show_email_modal = None
                    st.rerun()
        
        # KPI METRICS
        st.markdown("---")
        st.markdown("#### 📈 Performance Metrics")
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric("Total Students", stats.get('total_students', 0))
        with col2:
            st.metric("Average Marks", f"{stats.get('average_marks', 0):.1f}")
        with col3:
            st.metric("Pass Rate", f"{stats.get('pass_percentage', 0):.1f}%")
        with col4:
            st.metric("Average SGPA", f"{stats.get('average_sgpa', 0):.2f}")
        
        # TABS
        st.markdown("---")
        tab1, tab2, tab3, tab4 = st.tabs([
            "📊 Marks Distribution", 
            "🎯 CO Attainment", 
            "📐 PO Attainment", 
            "👥 Student List"
        ])
        
        with tab1:
            st.markdown("### 📊 Marks Distribution")
            marks = [s.get('total_marks', 0) for s in students.values()]
            fig_hist = px.histogram(
                x=marks, nbins=10, title="Marks Distribution",
                labels={'x': 'Total Marks', 'y': 'Number of Students'},
                color_discrete_sequence=['#667eea']
            )
            fig_hist.update_layout(template='plotly_white', height=400)
            st.plotly_chart(fig_hist, use_container_width=True)
            
            col_d1, col_d2 = st.columns(2)
            with col_d1:
                st.metric("Highest Marks", f"{stats.get('highest_marks', 0):.1f}")
                st.metric("Lowest Marks", f"{stats.get('lowest_marks', 0):.1f}")
            with col_d2:
                st.metric("Std Deviation", f"{stats.get('std_deviation', 0):.2f}")
                st.metric("Grade Range", f"A+ to F")
        
        with tab2:
            st.markdown("### 🎯 CO Attainment (Course Average)")
            st.info("📌 These are course averages. Individual student CO attainment varies.")
            
            co_attainment = course_data.get('co_attainment', {})
            if co_attainment:
                cos = list(co_attainment.keys())
                vals = list(co_attainment.values())
                fig_co = go.Figure(data=[go.Bar(
                    x=cos, y=vals,
                    text=[f'{v:.1f}%' for v in vals],
                    textposition='auto',
                    marker_color=['#4CAF50' if v >= 80 else '#FFC107' if v >= 60 else '#F44336' for v in vals]
                )])
                fig_co.update_layout(template='plotly_white', height=400, yaxis_title="Percentage (%)")
                st.plotly_chart(fig_co, use_container_width=True)
            else:
                st.info("No CO attainment data available")
        
        with tab3:
            st.markdown("### 📐 PO Attainment (Course Average)")
            st.info(f"📌 Auto-detected {len(course_data.get('po_columns', []))} POs from the uploaded file")
            
            po_attainment = course_data.get('po_attainment', {})
            if po_attainment:
                pos = list(po_attainment.keys())
                po_vals = list(po_attainment.values())
                if len(pos) >= 3:
                    fig = create_spider_plot(po_vals, pos, "PO Attainment (Course Average)")
                    st.pyplot(fig)
                else:
                    fig_po = go.Figure(data=[go.Bar(
                        x=pos, y=po_vals,
                        text=[f'{v:.1f}%' for v in po_vals],
                        textposition='auto',
                        marker_color='#764ba2'
                    )])
                    fig_po.update_layout(template='plotly_white', height=400, yaxis_title="Percentage (%)")
                    st.plotly_chart(fig_po, use_container_width=True)
            else:
                st.info("No PO attainment data available")
        
        with tab4:
            st.markdown("### 👥 Student List")
            
            search_term = st.text_input("🔍 Search by Name or ID", placeholder="Type to filter...")
            
            sorted_students = sorted(students.items(), key=lambda x: x[1].get('total_marks', 0), reverse=True)
            
            if search_term:
                sorted_students = [(sid, s) for sid, s in sorted_students 
                                  if search_term.lower() in sid.lower() or 
                                  search_term.lower() in s.get('name', '').lower()]
            
            student_table_data = []
            for rank, (sid, s) in enumerate(sorted_students, 1):
                student_table_data.append({
                    'Rank': rank,
                    'Student ID': sid,
                    'Name': s.get('name', 'N/A'),
                    'Total Marks': f"{s.get('total_marks', 0):.1f}",
                    'Grade': s.get('grade', 'N/A'),
                    'SGPA': f"{s.get('sgpa', 0):.2f}",
                    'Status': s.get('status', 'N/A')
                })
            
            if student_table_data:
                df_students = pd.DataFrame(student_table_data)
                st.dataframe(df_students, use_container_width=True, height=500, hide_index=True)
                
                csv = df_students.to_csv(index=False)
                st.download_button(
                    label="📥 Download Student List (CSV)",
                    data=csv,
                    file_name=f"student_list_{course_data.get('course_code', 'COURSE')}.csv",
                    mime="text/csv",
                    use_container_width=True
                )
            else:
                st.info("No students match your search")
    
    st.markdown('</div>', unsafe_allow_html=True)


def show_student_analytics():
    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown("## 📈 Student Analytics Dashboard")
    
    student_id = ""
    student_name = "Student"
    
    if st.session_state.user_type == "student":
        student_id = st.session_state.user_data.get('student_id', '')
        student_name = st.session_state.user_data.get('full_name', 'Student')
    elif st.session_state.user_type == "parent":
        student_id = st.session_state.user_data.get('student_linked', '')
        users = load_users()
        for s_info in users.get('students', {}).values():
            if s_info.get('student_id') == student_id:
                student_name = s_info.get('full_name', 'Your Child')
                break
    else:
        all_courses = load_all_courses()
        all_ids = set()
        for cd in all_courses.values():
            all_ids.update(cd.get('students', {}).keys())
        
        if all_ids:
            student_id = st.selectbox("Select Student to View Analytics", sorted(list(all_ids)))
    
    if not student_id:
        st.warning("No student ID found. Please ensure student accounts are properly set up.")
        st.markdown("</div>", unsafe_allow_html=True)
        return
    
    student_data = load_student_data(student_id)
    
    if not student_data:
        st.info("No academic data found for this student. Please upload course data first.")
        st.markdown("</div>", unsafe_allow_html=True)
        return
    
    latest_course_key = sorted(student_data.keys())[-1] if student_data else None
    latest_data = student_data.get(latest_course_key, {}) if latest_course_key else {}
    latest_student = latest_data.get('student_data', {})
    co_attainment = latest_data.get('co_attainment', {})
    po_attainment = latest_data.get('po_attainment', {})
    
    st.markdown(f"**👤 Viewing analytics for:** {student_name} (ID: {student_id})")
    
    tab1, tab2, tab3 = st.tabs([
        "📚 Course Results",
        "🎯 CO Attainment", 
        "📐 PO Attainment"
    ])
    
    with tab1:
        st.markdown("### 📚 Semester-wise Course Results")
        course_list = []
        for key, data in student_data.items():
            sd = data.get('student_data', {})
            course_list.append({
                'Course Code': data.get('course_code', ''),
                'Semester': data.get('semester', ''),
                'Total Marks': f"{sd.get('total_marks', 0):.1f}",
                'Grade': sd.get('grade', 'N/A'),
                'SGPA': f"{sd.get('sgpa', 0):.2f}",
                'Status': sd.get('status', 'N/A')
            })
        if course_list:
            df = pd.DataFrame(course_list)
            st.dataframe(df, use_container_width=True, hide_index=True)
            
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("Average Marks", f"{np.mean([float(c['Total Marks']) for c in course_list]):.1f}")
            with col2:
                st.metric("Average SGPA", f"{np.mean([float(c['SGPA']) for c in course_list]):.2f}")
            with col3:
                passed = len([c for c in course_list if c['Status'] == 'Pass'])
                st.metric("Courses Passed", f"{passed}/{len(course_list)}")
    
    with tab2:
        st.markdown("### 🎯 Course Outcome (CO) Attainment")
        st.info("📌 These are **your individual CO attainment percentages** for each course.")
        
        for key, data in student_data.items():
            co_pct = data.get('student_data', {}).get('co_attainment_pct', {})
            if co_pct:
                course_code = data.get('course_code', key)
                st.markdown(f"**📘 {course_code}**")
                cos = list(co_pct.keys())
                vals = list(co_pct.values())
                colors_bar = ['#4CAF50' if v >= 80 else '#FFC107' if v >= 60 else '#F44336' for v in vals]
                
                fig = go.Figure(data=[go.Bar(x=cos, y=vals, marker_color=colors_bar,
                                            text=[f'{v:.1f}%' for v in vals], textposition='auto')])
                fig.update_layout(title=f"CO Attainment - {course_code}", template='plotly_white',
                                yaxis_title="Percentage (%)", height=300)
                st.plotly_chart(fig, use_container_width=True)
    
    with tab3:
        st.markdown("### 📐 Program Outcome (PO) Attainment")
        st.info(f"📌 This shows **your individual PO attainment** based on your CO scores. Each student's PO attainment is unique.")
        
        for key, data in student_data.items():
            po_scores = data.get('po_attainment', {})
            student_po = data.get('student_data', {}).get('po_scores', {})
            
            display_po = student_po if student_po else po_scores
            
            if display_po:
                course_code = data.get('course_code', key)
                st.markdown(f"**📘 {course_code}**")
                
                pos = list(display_po.keys())
                vals = list(display_po.values())
                
                if len(pos) >= 3:
                    fig = create_spider_plot(vals, pos, f"PO Attainment - {course_code}")
                    st.pyplot(fig)
                else:
                    fig = go.Figure(data=[go.Bar(x=pos, y=vals, marker_color='#764ba2',
                                                text=[f'{v:.1f}%' for v in vals], textposition='auto')])
                    fig.update_layout(title=f"PO Attainment - {course_code}", template='plotly_white',
                                    yaxis_title="Percentage (%)", height=300)
                    st.plotly_chart(fig, use_container_width=True)
    
    st.markdown("</div>", unsafe_allow_html=True)


def show_teacher_panel():
    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown("## 👨‍🏫 Create User Account")
    
    tab1, tab2 = st.tabs(["👥 Create Student & Parent Account", "📋 Upload History"])
    
    with tab1:
        with st.form("create_student_parent_teacher"):
            st.markdown("#### Student Information")
            col1, col2 = st.columns(2)
            
            with col1:
                student_id = st.text_input("Student ID*", placeholder="EEE 078 07759")
                student_name = st.text_input("Student Full Name*")
                student_username = st.text_input("Student Username*", placeholder="Choose a username")
                student_password = st.text_input("Student Password*", type="password", placeholder="Choose a password")
            
            with col2:
                student_semester = st.text_input("Semester*", value=get_current_semester())
                student_email = st.text_input("Student Email (optional)")
            
            st.markdown("---")
            st.markdown("#### Parent/Guardian Information")
            col3, col4 = st.columns(2)
            
            with col3:
                parent_name = st.text_input("Parent Full Name*")
                parent_relationship = st.selectbox("Relationship*", ["Father", "Mother", "Guardian"])
                parent_username = st.text_input("Parent Username*", placeholder="Choose a username")
                parent_password = st.text_input("Parent Password*", type="password", placeholder="Choose a password")
            
            with col4:
                parent_email = st.text_input("Parent Email*")
                parent_contact = st.text_input("Parent Contact Number*")
            
            if st.form_submit_button("✅ Create Accounts", use_container_width=True, type="primary"):
                errors = []
                if not student_id: errors.append("Student ID required")
                if not student_name: errors.append("Student Name required")
                if not student_username: errors.append("Student Username required")
                if not student_password: errors.append("Student Password required")
                if not parent_name: errors.append("Parent Name required")
                if not parent_email: errors.append("Parent Email required")
                if not parent_username: errors.append("Parent Username required")
                if not parent_password: errors.append("Parent Password required")
                
                if errors:
                    for e in errors: st.error(f"❌ {e}")
                else:
                    users = load_users()
                    
                    if student_username in users.get('students', {}):
                        st.error(f"Student username '{student_username}' already exists!")
                    elif parent_username in users.get('parents', {}):
                        st.error(f"Parent username '{parent_username}' already exists!")
                    else:
                        users['students'][student_username] = {
                            "username": student_username,
                            "password": hash_password(student_password),
                            "email": student_email or parent_email,
                            "full_name": student_name,
                            "student_id": student_id,
                            "user_type": "student",
                            "is_active": True,
                            "semester": student_semester,
                            "parent_email": parent_email,
                            "parent_contact": parent_contact,
                            "parent_name": parent_name,
                            "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                            "created_by": st.session_state.username
                        }
                        
                        users['parents'][parent_username] = {
                            "username": parent_username,
                            "password": hash_password(parent_password),
                            "email": parent_email,
                            "full_name": f"{parent_name} ({parent_relationship})",
                            "student_linked": student_id,
                            "user_type": "parent",
                            "is_active": True,
                            "contact_no": parent_contact,
                            "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                            "created_by": st.session_state.username
                        }
                        
                        if save_users(users):
                            st.success(f"✅ Accounts created! Student: `{student_username}`, Parent: `{parent_username}`")
                            log_activity(st.session_state.username, "created_accounts", f"Student: {student_username}")
                            time.sleep(1.5)
                            st.rerun()
    
    with tab2:
        st.markdown("### 📋 Upload History")
        uploads = st.session_state.teacher_uploads.get(st.session_state.username, [])
        
        if uploads:
            upload_data = []
            for upload in reversed(uploads):
                upload_data.append({
                    'Date': upload.get('timestamp', ''),
                    'Semester': upload.get('semester', ''),
                    'Course Code': upload.get('course_code', ''),
                    'Students': upload.get('student_count', 0)
                })
            st.dataframe(pd.DataFrame(upload_data), use_container_width=True)
        else:
            st.info("ℹ️ No uploads yet")
    
    st.markdown("</div>", unsafe_allow_html=True)


def show_admin_panel():
    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown("## 👑 Admin Panel")
    
    tab1, tab2, tab3, tab4 = st.tabs(["👥 Create Accounts", "🔧 User Management", "💾 Data Management", "📜 Activity Log"])
    
    with tab1:
        col_a, col_b = st.columns(2)
        
        with col_a:
            st.markdown("### 👨‍🏫 Create Teacher")
            with st.form("admin_teacher"):
                t_user = st.text_input("Username*")
                t_pass = st.text_input("Password*", type="password")
                t_name = st.text_input("Full Name*")
                t_email = st.text_input("Email*")
                
                if st.form_submit_button("Create Teacher", use_container_width=True):
                    if all([t_user, t_pass, t_name, t_email]):
                        users = load_users()
                        if t_user not in users.get('teachers', {}):
                            users['teachers'][t_user] = {
                                "username": t_user,
                                "password": hash_password(t_pass),
                                "email": t_email,
                                "full_name": t_name,
                                "user_type": "teacher",
                                "is_active": True,
                                "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                            }
                            save_users(users)
                            st.success(f"✅ Teacher '{t_user}' created!")
                            log_activity(st.session_state.username, "created_teacher", t_user)
                        else:
                            st.error("Username exists")
        
        with col_b:
            st.markdown("### 👨‍🎓 Create Student & Parent")
            with st.form("admin_student"):
                st.markdown("**Student Information**")
                col_s1, col_s2 = st.columns(2)
                with col_s1:
                    s_id = st.text_input("Student ID*", placeholder="EEE 078 07759")
                    s_name = st.text_input("Student Full Name*")
                    s_username = st.text_input("Student Username*", placeholder="Choose username")
                    s_password = st.text_input("Student Password*", type="password", placeholder="Choose password")
                with col_s2:
                    s_email = st.text_input("Student Email")
                    s_sem = st.text_input("Semester*", value=get_current_semester())
                
                st.markdown("**Parent/Guardian Information**")
                col_p1, col_p2 = st.columns(2)
                with col_p1:
                    p_name = st.text_input("Parent Full Name*")
                    p_username = st.text_input("Parent Username*", placeholder="Choose username")
                    p_password = st.text_input("Parent Password*", type="password", placeholder="Choose password")
                with col_p2:
                    p_email = st.text_input("Parent Email*")
                    p_contact = st.text_input("Parent Contact No.")
                
                if st.form_submit_button("Create Accounts", use_container_width=True, type="primary"):
                    errors = []
                    if not s_id: errors.append("Student ID required")
                    if not s_name: errors.append("Student Name required")
                    if not s_username: errors.append("Student Username required")
                    if not s_password: errors.append("Student Password required")
                    if not p_name: errors.append("Parent Name required")
                    if not p_email: errors.append("Parent Email required")
                    if not p_username: errors.append("Parent Username required")
                    if not p_password: errors.append("Parent Password required")
                    
                    if errors:
                        for e in errors:
                            st.error(e)
                    else:
                        users = load_users()
                        
                        if s_username in users.get('students', {}):
                            st.error(f"Student username '{s_username}' already exists!")
                        elif p_username in users.get('parents', {}):
                            st.error(f"Parent username '{p_username}' already exists!")
                        elif s_username in users.get('teachers', {}):
                            st.error(f"Username '{s_username}' is already taken by a teacher!")
                        elif p_username in users.get('teachers', {}):
                            st.error(f"Username '{p_username}' is already taken by a teacher!")
                        else:
                            users['students'][s_username] = {
                                "username": s_username,
                                "password": hash_password(s_password),
                                "email": s_email or p_email,
                                "full_name": s_name,
                                "student_id": s_id,
                                "user_type": "student",
                                "is_active": True,
                                "semester": s_sem,
                                "parent_email": p_email,
                                "parent_contact": p_contact,
                                "parent_name": p_name,
                                "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                                "created_by": st.session_state.username
                            }
                            
                            users['parents'][p_username] = {
                                "username": p_username,
                                "password": hash_password(p_password),
                                "email": p_email,
                                "full_name": p_name,
                                "student_linked": s_id,
                                "user_type": "parent",
                                "is_active": True,
                                "contact_no": p_contact,
                                "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                                "created_by": st.session_state.username
                            }
                            
                            if save_users(users):
                                st.success(f"✅ Accounts created! Student: `{s_username}` | Parent: `{p_username}`")
                                log_activity(st.session_state.username, "created_accounts", f"Student: {s_username}, Parent: {p_username}")
                                time.sleep(1.5)
                                st.rerun()
    
    with tab2:
        st.markdown("### 🔧 User Management")
        users = load_users()
        user_type = st.selectbox("Filter by Type", ["teachers", "students", "parents", "admins"])
        
        if user_type in users:
            user_list = []
            for u, d in users[user_type].items():
                user_list.append({
                    'Username': u,
                    'Name': d.get('full_name', ''),
                    'Email': d.get('email', ''),
                    'Status': 'Active' if d.get('is_active', True) else 'Inactive',
                    'Created': d.get('created_at', 'N/A')
                })
            
            if user_list:
                df_users = pd.DataFrame(user_list)
                st.dataframe(df_users, use_container_width=True, hide_index=True)
                
                st.markdown("---")
                st.markdown("#### Toggle User Status")
                col_u1, col_u2 = st.columns(2)
                
                with col_u1:
                    selected_user = st.selectbox("Select User", list(users[user_type].keys()), key="user_select")
                
                with col_u2:
                    current_status = users[user_type][selected_user].get('is_active', True)
                    new_status = st.selectbox("Set Status", ["Active", "Inactive"], index=0 if current_status else 1)
                    
                    if st.button("Update Status", use_container_width=True):
                        users[user_type][selected_user]['is_active'] = (new_status == "Active")
                        save_users(users)
                        st.success(f"✅ Updated {selected_user} status to {new_status}")
                        log_activity(st.session_state.username, "user_status_change", f"{selected_user}: {new_status}")
                        st.rerun()
            else:
                st.info(f"No {user_type} found")
    
    with tab3:
        st.markdown("### 💾 Data Management")
        
        col_d1, col_d2 = st.columns(2)
        
        with col_d1:
            st.markdown("#### System Backup")
            if st.button("Create Full System Backup", use_container_width=True):
                backup_data = {
                    'users': load_users(),
                    'activity_log': st.session_state.activity_log,
                    'date': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    'backup_version': '1.0'
                }
                
                backup_json = json.dumps(backup_data, indent=4, default=str)
                st.download_button(
                    label="Download Backup File",
                    data=backup_json,
                    file_name=f"edutrack_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
                    mime="application/json",
                    use_container_width=True
                )
                log_activity(st.session_state.username, "created_backup", "Full system backup")
        
        with col_d2:
            st.markdown("#### Course Data Export")
            
            all_courses = load_all_courses()
            if all_courses:
                st.info(f"Total courses stored: {len(all_courses)}")
                
                export_data = {
                    'courses': {},
                    'export_date': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                }
                
                for key, course in all_courses.items():
                    export_data['courses'][key] = {
                        'semester': course.get('semester', ''),
                        'course_code': course.get('course_code', ''),
                        'course_name': course.get('course_name', ''),
                        'teacher': course.get('teacher', ''),
                        'course_stats': course.get('course_stats', {}),
                        'student_count': len(course.get('students', {}))
                    }
                
                export_json = json.dumps(export_data, indent=4, default=str)
                st.download_button(
                    label="Export Course Summary",
                    data=export_json,
                    file_name=f"course_summary_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
                    mime="application/json",
                    use_container_width=True
                )
            else:
                st.warning("No course data available")
        
        st.markdown("---")
        st.markdown("#### Reset Options")
        st.warning("⚠️ These actions are irreversible. Please create a backup before proceeding.")
        
        col_r1, col_r2, col_r3 = st.columns(3)
        
        with col_r1:
            if st.button("Reset Course Data", use_container_width=True):
                st.session_state.show_reset_confirm = "courses"
                st.rerun()
        
        with col_r2:
            if st.button("Reset Activity Log", use_container_width=True):
                st.session_state.show_reset_confirm = "activity"
                st.rerun()
        
        with col_r3:
            if st.button("Reset All Data", use_container_width=True, type="secondary"):
                st.session_state.show_reset_confirm = "all"
                st.rerun()
        
        if st.session_state.get('show_reset_confirm'):
            st.markdown("---")
            reset_type = st.session_state.show_reset_confirm
            
            if reset_type == "courses":
                st.error("⚠️ Are you sure you want to delete ALL course data?")
                st.markdown("This will remove all uploaded course results and student data.")
                
                col_confirm1, col_confirm2 = st.columns(2)
                with col_confirm1:
                    if st.button("Yes, Delete All Course Data", use_container_width=True):
                        try:
                            data_dir = Path("course_data")
                            if data_dir.exists():
                                for file in data_dir.glob("*"):
                                    file.unlink()
                            st.success("All course data has been deleted")
                            log_activity(st.session_state.username, "reset_data", "Deleted all course data")
                            st.session_state.show_reset_confirm = None
                            time.sleep(1.5)
                            st.rerun()
                        except Exception as e:
                            st.error(f"Error: {e}")
                
                with col_confirm2:
                    if st.button("Cancel", use_container_width=True):
                        st.session_state.show_reset_confirm = None
                        st.rerun()
            
            elif reset_type == "activity":
                st.error("⚠️ Are you sure you want to clear the activity log?")
                
                col_confirm1, col_confirm2 = st.columns(2)
                with col_confirm1:
                    if st.button("Yes, Clear Activity Log", use_container_width=True):
                        st.session_state.activity_log = []
                        st.success("Activity log cleared")
                        log_activity(st.session_state.username, "reset_data", "Cleared activity log")
                        st.session_state.show_reset_confirm = None
                        time.sleep(1.5)
                        st.rerun()
                
                with col_confirm2:
                    if st.button("Cancel", use_container_width=True):
                        st.session_state.show_reset_confirm = None
                        st.rerun()
            
            elif reset_type == "all":
                st.error("⚠️ WARNING: This will delete ALL data including courses, models, and logs!")
                
                col_confirm1, col_confirm2 = st.columns(2)
                with col_confirm1:
                    if st.button("Yes, Delete Everything", use_container_width=True, type="secondary"):
                        try:
                            data_dir = Path("course_data")
                            if data_dir.exists():
                                for file in data_dir.glob("*"):
                                    file.unlink()
                            
                            ml_dir = Path("ml_models")
                            if ml_dir.exists():
                                for file in ml_dir.glob("*"):
                                    file.unlink()
                            
                            st.session_state.activity_log = []
                            
                            st.success("All data has been reset")
                            log_activity(st.session_state.username, "reset_data", "Full system reset")
                            st.session_state.show_reset_confirm = None
                            time.sleep(1.5)
                            st.rerun()
                        except Exception as e:
                            st.error(f"Error: {e}")
                
                with col_confirm2:
                    if st.button("Cancel", use_container_width=True):
                        st.session_state.show_reset_confirm = None
                        st.rerun()
    
    with tab4:
        st.markdown("### 📜 Activity Log")
        st.markdown("Track all system activities and user actions")
        
        activity_log = st.session_state.activity_log
        
        if activity_log:
            col_al1, col_al2 = st.columns([3, 1])
            
            with col_al1:
                st.info(f"Total activities logged: {len(activity_log)}")
            
            with col_al2:
                if st.button("Refresh Log", use_container_width=True):
                    st.rerun()
            
            col_f1, col_f2 = st.columns(2)
            with col_f1:
                all_actions = list(set(log['action'] for log in activity_log))
                filter_action = st.selectbox("Filter by Action", ["All"] + all_actions)
            
            with col_f2:
                search_user = st.text_input("Search by Username", placeholder="Type username...")
            
            filtered_logs = activity_log.copy()
            if filter_action != "All":
                filtered_logs = [log for log in filtered_logs if log['action'] == filter_action]
            if search_user:
                filtered_logs = [log for log in filtered_logs if search_user.lower() in log['username'].lower()]
            
            log_data = []
            for log in reversed(filtered_logs):
                log_data.append({
                    'Timestamp': log.get('timestamp', 'N/A'),
                    'Username': log.get('username', 'N/A'),
                    'Action': log.get('action', 'N/A'),
                    'Details': log.get('details', '')[:50] + ('...' if len(log.get('details', '')) > 50 else '')
                })
            
            if log_data:
                df_logs = pd.DataFrame(log_data)
                st.dataframe(df_logs, use_container_width=True, hide_index=True, height=400)
                
                export_log = pd.DataFrame([{
                    'Timestamp': log.get('timestamp', ''),
                    'Username': log.get('username', ''),
                    'Action': log.get('action', ''),
                    'Details': log.get('details', '')
                } for log in reversed(activity_log)])
                
                csv_log = export_log.to_csv(index=False)
                st.download_button(
                    label="Download Activity Log (CSV)",
                    data=csv_log,
                    file_name=f"activity_log_{datetime.now().strftime('%Y%m%d')}.csv",
                    mime="text/csv",
                    use_container_width=True
                )
            else:
                st.info("No matching activities found")
        else:
            st.info("No activity logged yet. System activities will appear here as users interact with the platform.")
    
    st.markdown("</div>", unsafe_allow_html=True)


def show_sidebar():
    with st.sidebar:
        st.markdown("""
        <div style="text-align: center; margin-bottom: 1rem;">
            <h1 style="font-size: 3rem;">🎓</h1>
            <h3>EduTrack Pro 2026</h3>
        </div>
        """, unsafe_allow_html=True)
        
        st.markdown(f"""
        <div style="background: #667eea; color: white; padding: 1rem; border-radius: 10px; margin-bottom: 1rem;">
            <strong>{st.session_state.user_data.get('full_name', 'User')}</strong><br>
            <small>{st.session_state.user_type.title()}</small>
        </div>
        """, unsafe_allow_html=True)
        
        st.markdown("### 🧭 Navigation")
        
        if st.session_state.user_type in ["teacher", "admin"]:
            if st.button("📤 Upload Data", use_container_width=True):
                st.session_state.current_page = "upload"
                st.rerun()
            if st.button("📊 Course Reports", use_container_width=True):
                st.session_state.current_page = "reports"
                st.rerun()
        
        if st.button("📈 Student Analytics", use_container_width=True):
            st.session_state.current_page = "student_analytics"
            st.rerun()
        
        if st.session_state.user_type == "teacher":
            if st.button("👥 Create User Account", use_container_width=True):
                st.session_state.current_page = "teacher_panel"
                st.rerun()
        
        if st.session_state.user_type == "admin":
            if st.button("👑 Admin Panel", use_container_width=True):
                st.session_state.current_page = "admin_panel"
                st.rerun()
        
        st.markdown("---")
        
        # Change Password Button (for all users except maybe guests)
        if st.button("🔐 Change Password", use_container_width=True):
            st.session_state.show_change_password = True
            st.session_state.current_page = "change_password"
            st.rerun()
        
        if st.button("🚪 Logout", use_container_width=True, type="secondary"):
            for key in ['logged_in', 'user_type', 'username', 'user_data']:
                st.session_state[key] = False if key == 'logged_in' else ""
            st.session_state.current_page = "login"
            st.session_state.show_change_password = False
            st.rerun()


# MAIN APP
def main():
    apply_professional_theme()
    
    if not st.session_state.logged_in:
        show_login_page()
        return
    
    # Handle change password page
    if st.session_state.get('show_change_password', False) or st.session_state.current_page == "change_password":
        show_change_password_page()
        show_sidebar()
        return
    
    show_sidebar()
    
    page = st.session_state.current_page
    
    if page == "upload" and st.session_state.user_type in ["teacher", "admin"]:
        show_upload_page()
    elif page == "reports" and st.session_state.user_type in ["teacher", "admin"]:
        show_course_reports()
    elif page == "student_analytics":
        show_student_analytics()
    elif page == "teacher_panel" and st.session_state.user_type == "teacher":
        show_teacher_panel()
    elif page == "admin_panel" and st.session_state.user_type == "admin":
        show_admin_panel()
    else:
        show_upload_page() if st.session_state.user_type in ["teacher", "admin"] else show_student_analytics()
    
    st.markdown("---")
    st.markdown("""
    <div style="text-align: center; padding: 10px; color: #666;">
        <p>EduTrack Pro 2026 | Department of EEE | Stamford University Bangladesh</p>
    </div>
    """, unsafe_allow_html=True)


if __name__ == "__main__":
    # Import pickle for data persistence
    import pickle
    Path("course_data").mkdir(exist_ok=True)
    main()
